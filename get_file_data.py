import datetime
import re
import xml.dom.minidom
from typing import List, Optional

import openpyxl
import sqlalchemy.orm

import auxiliary
import configuration
import db_calls
import models
import process_emails
import response_models
import tmdb_calls
import utilities

unordered_words = ['the', 'a', 'an', 'i', 'un', 'le', 'la', 'les', 'um', 'o', 'el', 'as', 'os']


class InsertionResult:
    """To store the results of an insertion from a file."""

    start_datetime: datetime.datetime
    end_datetime: datetime.datetime
    total_nb_sessions_in_file: int
    nb_updated_sessions: int
    nb_added_sessions: int
    nb_deleted_sessions: int
    nb_new_shows: int

    def __init__(self):
        self.total_nb_sessions_in_file = 0
        self.nb_updated_sessions = 0
        self.nb_added_sessions = 0
        self.nb_deleted_sessions = 0
        self.nb_new_shows = 0


class ChannelInsertion:
    channels: str


class Cinemundo(ChannelInsertion):
    channels = ['Cinemundo']

    @staticmethod
    def process_title(title: str) -> [str, bool, int]:
        """
        Process the title, removing special markers:
        - VP - for portuguese audio.

        :param title: the title as is in the file.
        :return: a tuple with the clean title, whether or not it is a session with the portuguese voice and the season, when applicable.
        """

        # Replace all quotation marks for the same quotation mark
        title = re.sub('[´`]', '\'', title)

        # Search for VP in the end of the title
        vp = title.endswith('VP')

        if vp:
            # Remove the VP when it exists
            title = title[:-3]

        # Sometimes the title contains multiple titles
        char_pos = auxiliary.search_chars(title, ['/'])

        # Get the first title
        if len(char_pos[0]) > 0:
            title = title[:char_pos[0][0]]

        # Check if it is a series
        series = re.search(r'S[0-9]+', title.strip())

        if series is not None:
            season = int(series.group(0)[1:])
            title = title[:series.span(0)[0]]
        else:
            season = None

        return title.strip(), vp, season

    @staticmethod
    def add_file_data(db_session: sqlalchemy.orm.Session, filename: str) -> Optional[InsertionResult]:
        """
        Add the data, in the file, to the DB.

        :param db_session: the DB session.
        :param filename: the path to the file.
        :return: the InsertionResult.
        """

        wb = openpyxl.load_workbook(filename)

        first_event_datetime = None

        insertion_result = InsertionResult()

        # Skip row 1, with the headers
        for row in wb.active.iter_rows(min_row=3, max_col=12):
            # Skip rows that do not contain a date
            if row[0].value is None:
                continue

            # Get the data
            date = datetime.datetime.strptime(str(row[0].value), '%Y%m%d')
            time = row[1].value
            original_title = str(row[2].value)
            localized_title = str(row[3].value)
            synopsis = row[4].value
            year = int(row[5].value)
            age_classification = row[6].value
            directors = row[7].value
            cast = row[8].value
            subgenre = row[9].value  # Obtained in portuguese

            # Combine the date with the time
            date_time = date.replace(hour=time.hour, minute=time.minute)

            # Get the first event's datetime
            if first_event_datetime is None:
                first_event_datetime = date_time

            # Process the titles
            localized_title, vp, _ = Cinemundo.process_title(localized_title)
            audio_language = 'pt' if vp else None

            original_title, _, season = Cinemundo.process_title(original_title)

            if season is not None:
                is_movie = False
                genre = 'Series'
            else:
                is_movie = True
                genre = 'Movie'

            # Process the directors
            if directors is not None:
                directors = re.split(',| e ', directors)

            # Get the channel's id
            channel_id = db_calls.get_channel_name(db_session, 'Cinemundo').id

            # Process an entry
            insertion_result = process_file_entry(db_session, insertion_result, original_title, localized_title,
                                                  is_movie, genre, date_time, channel_id, year, directors, subgenre,
                                                  synopsis, season, None, cast=cast,
                                                  age_classification=age_classification, audio_languages=audio_language)

            if insertion_result is None:
                return None

        if insertion_result.total_nb_sessions_in_file != 0:
            db_calls.commit(db_session)

            # Delete old sessions for the same time period
            file_start_datetime = first_event_datetime - datetime.timedelta(minutes=5)
            file_end_datetime = date_time + datetime.timedelta(minutes=5)

            nb_deleted_sessions = delete_old_sessions(db_session, file_start_datetime, file_end_datetime,
                                                      Cinemundo.channels)

            # Set the remaining information
            insertion_result.nb_deleted_sessions = nb_deleted_sessions
            insertion_result.start_datetime = file_start_datetime
            insertion_result.end_datetime = file_end_datetime

            return insertion_result
        else:
            return None


class Odisseia(ChannelInsertion):
    channels = ['Odisseia']

    @staticmethod
    def process_title(title: str) -> str:
        """ Process the title, removing special markers and reformatting the title. """

        # Replace all quotation marks for the same quotation mark
        return re.sub('[´`]', '\'', title)

    @staticmethod
    def add_file_data(db_session: sqlalchemy.orm.Session, filename: str) -> Optional[InsertionResult]:
        """
        Add the data, in the file, to the DB.

        :param db_session: the DB session.
        :param filename: the path to the file.
        :return: the InsertionResult.
        """

        dom_tree = xml.dom.minidom.parse(filename)
        collection = dom_tree.documentElement

        # Get all events
        events = collection.getElementsByTagName('Event')

        # If there are no events
        if len(events) == 0:
            return None

        first_event_datetime = None

        insertion_result = InsertionResult()

        # Process each event
        for event in events:
            # --- START DATA GATHERING ---
            # Get the date and time
            begin_time = event.getAttribute('beginTime')
            date_time = datetime.datetime.strptime(begin_time, '%Y%m%d%H%M%S')

            # Get the first event's datetime
            if first_event_datetime is None:
                first_event_datetime = date_time

            # Get the event's duration in minutes
            duration = int(int(event.getAttribute('duration')) / 60)

            # Inside the Event -> EpgProduction
            epg_production = event.getElementsByTagName('EpgProduction')[0]

            # Get the genre
            genre_list = epg_production.getElementsByTagName('Genere')

            # Check if it is the genre that we are assuming it always is
            if len(genre_list) > 0 and 'Document' not in genre_list[0].firstChild.nodeValue:
                print_message('not a documentary', True, str(event.getAttribute('beginTime')))

            # Subgenre is in portuguese
            subgenre = epg_production.getElementsByTagName('Subgenere')[0].firstChild.nodeValue

            # Age classification
            age_classification = epg_production.getElementsByTagName('ParentalRating')[0].firstChild.nodeValue

            # Inside the Event -> EpgProduction -> EpgText
            epg_text = epg_production.getElementsByTagName('EpgText')[0]

            # Get the localized title, in this case the portuguese one
            localized_title = epg_text.getElementsByTagName('Name')[0].firstChild.nodeValue

            # Get the localized synopsis, in this case the portuguese one
            short_description = epg_text.getElementsByTagName('ShortDescription')

            if short_description is not None and short_description[0].firstChild is not None:
                synopsis = short_description[0].firstChild.nodeValue
            else:
                synopsis = None

            # Iterate over the ExtendedInfo elements
            extended_info_elements = epg_text.getElementsByTagName('ExtendedInfo')

            original_title = None
            directors = None
            season = None
            episode = None
            year = None
            countries = None
            cast = None

            for extended_info in extended_info_elements:
                attribute = extended_info.getAttribute('name')

                if attribute == 'OriginalEventName' and extended_info.firstChild is not None:
                    original_title = extended_info.firstChild.nodeValue
                elif attribute == 'Year' and extended_info.firstChild is not None:
                    year = int(extended_info.firstChild.nodeValue)

                    # Sometimes the year is 0
                    if year == 0:
                        year = None
                elif attribute == 'Director' and extended_info.firstChild is not None:
                    directors = extended_info.firstChild.nodeValue
                elif attribute == 'Casting' and extended_info.firstChild is not None:
                    cast = extended_info.firstChild.nodeValue
                elif attribute == 'Nationality' and extended_info.firstChild is not None:
                    countries = extended_info.firstChild.nodeValue
                elif attribute == 'Cycle' and extended_info.firstChild is not None:
                    season = int(extended_info.firstChild.nodeValue)
                elif attribute == 'EpisodeNumber' and extended_info.firstChild is not None:
                    episode = int(extended_info.firstChild.nodeValue)

            # Get the channel's id
            channel_id = db_calls.get_channel_name(db_session, 'Odisseia').id

            # Process titles
            original_title = Odisseia.process_title(original_title)
            localized_title = Odisseia.process_title(localized_title)

            # Process the directors
            if directors is not None:
                directors = directors.split(',')

            is_movie = season is None
            genre = 'Documentary'

            # --- END DATA GATHERING ---

            # Process file entry
            insertion_result = process_file_entry(db_session, insertion_result, original_title, localized_title,
                                                  is_movie, genre, date_time, channel_id, year, directors, subgenre,
                                                  synopsis, season, episode, cast=cast, duration=duration,
                                                  countries=countries, age_classification=age_classification)

            if insertion_result is None:
                return None

        db_calls.commit(db_session)

        # Delete old sessions for the same time period
        file_start_datetime = first_event_datetime - datetime.timedelta(minutes=5)
        file_end_datetime = date_time + datetime.timedelta(minutes=5)

        nb_deleted_sessions = delete_old_sessions(db_session, file_start_datetime, file_end_datetime, Odisseia.channels)

        # Set the remaining information
        insertion_result.nb_deleted_sessions = nb_deleted_sessions
        insertion_result.start_datetime = file_start_datetime
        insertion_result.end_datetime = file_end_datetime

        return insertion_result


class TVCine(ChannelInsertion):
    channels = ['TVCine Top', 'TVCine Edition', 'TVCine Emotion', 'TVCine Action']

    @staticmethod
    def fix_title_order(title: str):
        """
        Fix the unordered parts of the title, if any.

        :param title: the title as is in the file.
        :return: the final title.
        """

        building = False
        start = 0

        for i in range(len(title)):
            if title[i] == ',':
                building = True
                start = i
            elif building:
                if re.match('[0-9a-zA-Z ]', title[i]) is None:
                    building = False

                    if title[start + 1:i].strip().lower() in unordered_words:
                        title = (' '.join([title[start + 1:i].strip(), title[:start].strip(), title[i:].strip()]))
                        break

        if building and title[start + 1:].strip().lower() in unordered_words:
            title = (' '.join([title[start + 1:].strip(), title[:start].strip()]))

        return title

    @staticmethod
    def process_title(title: str) -> [str, bool, bool]:
        """
        Process the title, removing special markers and reformatting the title:
        - (VP) - for portuguese audio;
        - (VO) - for original audio;
        - (extended cut) or (versão alargada) - for extended cut;
        - some of the words come unordered, thus the unordered list of words.

        :param title: the title as is in the file.
        :return: a tuple with the clean title, whether or not it is a session with the portuguese voice and whether or
        not it is a session with the extended cut.
        """

        # Replace all quotation marks for the same quotation mark
        title = re.sub('[´`]', '\'', title)

        search_result = auxiliary.search_chars(title, ['(', ')'])
        vp = False
        extended_cut = False

        # If the number of opening parenthesis is not the same as the closing ones
        if len(search_result[0]) != len(search_result[1]):
            return title.strip(), vp

        # From the last position of the parenthesis
        for i in range(len(search_result[0]) - 1, -1, -1):
            start_pos = search_result[0][i]
            end_pos = search_result[1][i]

            text = title[start_pos + 1:end_pos]

            if text == 'VP':
                vp = True
            elif text == 'VO':
                vp = False
            elif text == 'versão alargada' or text == 'extended cut':
                extended_cut = True
            elif re.search(r'[0-9]{4}', text.strip()):
                pass
            else:
                continue

            # Remove the parenthesis and its context
            title = title[:search_result[0][i]] + title[search_result[1][i] + 1:]

        return TVCine.fix_title_order(title).strip(), vp, extended_cut

    @staticmethod
    def add_file_data(db_session: sqlalchemy.orm.Session, filename: str) -> Optional[InsertionResult]:
        """
        Add the data, in the file, to the DB.

        :param db_session: the DB session.
        :param filename: the path to the file.
        :return: the InsertionResult.
        """

        wb = openpyxl.load_workbook(filename)

        insertion_result = InsertionResult()

        first_event_datetime = None

        # Skip row 1, with the headers
        for row in wb.active.iter_rows(min_row=2, max_col=15):
            # Skip rows that contain only the date
            if row[0].value is None:
                continue

            # Skip the rows in which the year is not a number (header rows)
            if not isinstance(row[4].value, int):
                continue

            # Get the data
            channel_name = row[0].value
            date = row[1].value
            time = row[2].value
            original_title = str(row[3].value)
            year = int(row[4].value)
            age_classification = row[5].value
            genre = row[6].value
            duration = int(row[7].value)
            languages = row[8].value
            countries = row[9].value
            synopsis = row[10].value
            directors = row[11].value
            cast = row[12].value
            localized_title = str(row[13].value)
            # episode_title = row[14].value

            # Combine the date with the time
            date_time = date.replace(hour=time.hour, minute=time.minute)

            # Get the first event's datetime
            if first_event_datetime is None:
                first_event_datetime = date_time

            # Check if it matches the regex of a series
            series = re.search('(.+) T([0-9]+),[ ]+([0-9]+)', localized_title.strip())

            # If it is a series, extract it's season and episode
            if series:
                localized_title = series.group(1)
                is_movie = False

                season = int(series.group(2))
                episode = int(series.group(3))

                # episode_synopsis = synopsis
                synopsis = None

                # Also get the original title without the season and episode
                series = re.search('(.+) T([0-9]+),[ ]+([0-9]+)', original_title.strip())

                if series:
                    original_title = series.group(1)
            else:
                season = None
                episode = None

                is_movie = True

            # Process the titles
            localized_title, vp, extended_cut = TVCine.process_title(localized_title)
            audio_language = 'pt' if vp else None

            original_title, _, _ = TVCine.process_title(original_title)

            # Sometimes the cast is switched with the director
            if cast is not None and directors is not None:
                cast_commas = auxiliary.search_chars(cast, [','])[0]
                director_commas = auxiliary.search_chars(directors, [','])[0]

                # When that happens, switch them
                if len(cast_commas) < len(director_commas):
                    aux = cast
                    cast = directors
                    directors = aux

            # Process the directors
            if directors is not None:
                directors = directors.split(',')

            # Genre is movie, series, documentary, news...
            if 'Document' not in genre:
                subgenre = genre  # Subgenre is in portuguese
                genre = 'Movie' if is_movie else 'Series'
            else:
                genre = 'Documentary'
                subgenre = None

            channel_name = 'TVCine ' + channel_name.strip().split()[1]
            channel_id = db_calls.get_channel_name(db_session, channel_name).id

            # Process file entry
            insertion_result = process_file_entry(db_session, insertion_result, original_title, localized_title,
                                                  is_movie, genre, date_time, channel_id, year, directors, subgenre,
                                                  synopsis, season, episode, cast=cast, duration=duration,
                                                  countries=countries, age_classification=age_classification,
                                                  audio_languages=languages, session_audio_language=audio_language,
                                                  extended_cut=extended_cut)

            if insertion_result is None:
                return None

        if insertion_result.total_nb_sessions_in_file != 0:
            db_calls.commit(db_session)

            # Delete old sessions for the same time period
            file_start_datetime = first_event_datetime - datetime.timedelta(minutes=5)
            file_end_datetime = date_time + datetime.timedelta(minutes=5)

            nb_deleted_sessions = delete_old_sessions(db_session, file_start_datetime, file_end_datetime,
                                                      TVCine.channels)

            # Set the remaining information
            insertion_result.nb_deleted_sessions = nb_deleted_sessions
            insertion_result.start_datetime = file_start_datetime
            insertion_result.end_datetime = file_end_datetime

            return insertion_result
        else:
            return None


channel_insertion_list = [Cinemundo, Odisseia, TVCine]


def print_message(message: str, warning: bool, identification: str):
    """
    Print a message for a warning or an error in a show.

    :param message: the message.
    :param warning: True, if it is a warning. False, for an error.
    :param identification: a string that identifies the show in the context of the message.
    """

    message_type = 'Warning' if warning else 'Error'

    print('%s: The show |%s| - %s!' % (message_type, identification, message))


def process_file_entry(db_session: sqlalchemy.orm.Session, insertion_result: InsertionResult, original_title: str,
                       localized_title: str, is_movie: bool, genre: str, date_time: datetime.datetime, channel_id: int,
                       year: Optional[int], directors: Optional[List[str]], subgenre: Optional[str],
                       synopsis: Optional[str], season: Optional[int], episode: Optional[int],
                       cast: Optional[str] = None, duration: Optional[int] = None, audio_languages: str = None,
                       countries: str = None, age_classification: Optional[str] = None,
                       session_audio_language: Optional[str] = None, extended_cut: bool = False) \
        -> Optional[InsertionResult]:
    """
    Process an entry in the file, inserting all of the needed data.

    :param db_session: the db session.
    :param insertion_result: the insertion result.
    :param original_title: the original title.
    :param localized_title: the localized title.
    :param is_movie: whether or not it is a movie.
    :param genre: the type of show (movie, series, documentary, ...).
    :param date_time: the datetime.
    :param channel_id: the id of the channel.
    :param year: the year of the show.
    :param directors: the directors of the show.
    :param subgenre: the subgenre of the show.
    :param synopsis: the synopsis.
    :param season: the season.
    :param episode: the episode.
    :param cast: the cast of the show.
    :param duration: the duration.
    :param audio_languages: the languages of the audio.
    :param countries: the countries.
    :param age_classification: the age classification.
    :param session_audio_language: the audio language for the current session.
    :param extended_cut: whether or not this is the extended cut.
    :return: the updated insertion result, or None if there's a fatal error.
    """

    new_show = False

    # Search the ChannelShowDataCorrection
    channel_show_data = db_calls.search_channel_show_data_correction(db_session, channel_id, is_movie, original_title,
                                                                     localized_title, directors=directors, year=year,
                                                                     subgenre=subgenre)

    # If no match was found
    if channel_show_data is None:
        # Insert the ShowData, if necessary
        new_show, show_data = db_calls.insert_if_missing_show_data(db_session, localized_title, cast=cast,
                                                                   original_title=original_title, duration=duration,
                                                                   synopsis=synopsis, year=year, genre=genre,
                                                                   subgenre=subgenre, audio_languages=audio_languages,
                                                                   countries=countries, directors=directors,
                                                                   age_classification=age_classification,
                                                                   is_movie=is_movie, season=season)

        # If it is a new show, search the TMDB
        if new_show:
            insertion_result.nb_new_shows += 1
            tmdb_show = search_tmdb_match(db_session, show_data)

            # If it found a match in TMDB
            if tmdb_show:
                tmdb_show_data = db_calls.get_show_data_tmdb_id(db_session, tmdb_show.id)

                # If an entry with that TMDB id already exists, delete the new one
                if tmdb_show_data is not None:
                    db_session.delete(show_data)
                    show_data = tmdb_show_data

                # If not, update the information
                else:
                    utilities.update_show_data_with_tmdb(show_data, tmdb_show)

                # If there are differences between the data from the TMDB and the one in the file
                if show_data.original_title.casefold() != original_title.casefold() \
                        or (year is not None and show_data.year != year):
                    db_calls.register_channel_show_data_correction(db_session, channel_id, show_data.id, is_movie,
                                                                   original_title, localized_title,
                                                                   directors=directors, year=year, subgenre=subgenre)

    # If it found a matching ChannelShowData
    else:
        show_data = db_calls.get_show_data_id(db_session, channel_show_data.show_id)

    # Process a show session
    return process_show_session(db_session, insertion_result, show_data, new_show, season, episode,
                                date_time, channel_id, audio_language=session_audio_language, extended_cut=extended_cut)


def process_show_session(db_session: sqlalchemy.orm.Session, insertion_result: InsertionResult,
                         show_data: models.ShowData, new_show: bool, season: Optional[int], episode: Optional[int],
                         date_time: datetime.datetime, channel_id: int, audio_language: str = None,
                         extended_cut: bool = False) -> Optional[InsertionResult]:
    """
    Process a show session.

    :param db_session: the db session.
    :param insertion_result: the insertion result.
    :param show_data: the corresponding show data.
    :param new_show: whether or not the show data was new.
    :param season: the season.
    :param episode: the episode.
    :param date_time: the datetime.
    :param channel_id: the id of the channel.
    :param audio_language: the audio language.
    :param extended_cut: whether or not this is the extended cut.
    :return: the updated insertion result, or None if there's a fatal error.
    """

    if show_data is None:
        print_message('insertion of Show Data', False, str(date_time))
        return None

    if show_data.is_movie and show_data.director is None:
        print_message('director not provided', True, str(date_time))

    insertion_result.total_nb_sessions_in_file += 1

    # If it is a new show
    if new_show:
        add_show = True
    else:
        existing_show_session = db_calls.search_existing_session(db_session, season, episode, date_time, channel_id,
                                                                 show_data.id)

        # If there's already an existing session
        if existing_show_session is not None:
            add_show = False

            existing_show_session.date_time = date_time
            existing_show_session.update_timestamp = datetime.datetime.now()
        else:
            add_show = True

    # Insert the new session
    if add_show:
        show_session = db_calls.register_show_session(db_session, season, episode, date_time, channel_id, show_data.id,
                                                      audio_language=audio_language, extended_cut=extended_cut,
                                                      should_commit=False)

        if show_session is None:
            print('Session insertion failed!')
            return insertion_result

        insertion_result.nb_added_sessions += 1
    else:
        insertion_result.nb_updated_sessions += 1

    return insertion_result


def delete_old_sessions(db_session: sqlalchemy.orm.Session, start_datetime: datetime.datetime,
                        end_datetime: datetime.datetime, channels: List[str]) -> int:
    """
    Delete sessions that no longer exist.
    Send emails to the users whose reminders are associated with such sessions.

    :param db_session: the DB session.
    :param start_datetime: the start of the interval of interest.
    :param end_datetime: the end of the interval of interest.
    :param channels: the set of channels.
    :return: the number of deleted sessions.
    """

    nb_deleted_sessions = 0

    # Get the old show sessions
    old_sessions = db_calls.search_old_sessions(db_session, start_datetime, end_datetime, channels)

    for s in old_sessions:
        nb_deleted_sessions += 1

        # Get the reminders associated with this session
        reminders = db_calls.get_reminders_session(db_session, s.id)

        if len(reminders) != 0:
            # Get the session
            show_session = db_calls.get_show_session_complete(db_session, s.id)
            show_result = response_models.LocalShowResult.create_from_show_session(show_session[0], show_session[1],
                                                                                   show_session[2])

            # Warn all users with the reminders for this session
            for r in reminders:
                user = db_calls.get_user_id(db_session, r.user_id)

                process_emails.send_deleted_sessions_email(user.email, [show_result])

                # Delete the reminder
                db_session.delete(r)

        # Delete the session
        db_session.delete(s)

    db_session.commit()

    return nb_deleted_sessions


def search_tmdb_match(db_session: sqlalchemy.orm.Session, show_data: models.ShowData, use_year: bool = True) \
        -> Optional[tmdb_calls.TmdbShow]:
    """
    Search for a TMDB match.

    :param db_session: the DB session.
    :param show_data: the data of the show.
    :param use_year: whether to use the year in the search or not.
    :return: the TMDB show that matches.
    """

    if use_year:
        year = show_data.year
    else:
        year = None

    _, tmdb_shows = tmdb_calls.search_shows_by_text(db_session, show_data.original_title, is_movie=show_data.is_movie,
                                                    year=year)

    # Starting score
    if year is not None:
        starting_score = 5
    else:
        starting_score = 0

    best = None
    best_score = 0

    # Iterate over the results from the
    for t in tmdb_shows:
        score = starting_score

        # If the genre is a match
        if show_data.genre != 'Movie' and show_data.genre != 'Series' and show_data.genre in t.genres:
            score += 5

        # If the title is an exact match
        if t.original_title.casefold() == show_data.original_title.casefold():
            score += 20

        # Otherwise search for the director
        if show_data.director is not None and score < 25:
            directors = show_data.director.split(',')

            crew_list = tmdb_calls.get_show_crew_members(t.id, t.is_movie)

            found_director = False

            for member in crew_list:
                # Check the director's name in a case insensitive manner
                if member.name.casefold() in map(str.casefold, directors):
                    for job in member.jobs:
                        if job == 'Director':
                            score += 20
                            found_director = True
                            break

                    if found_director:
                        break

        # Update the best match
        if score > best_score:
            best = t
            best_score = score

            # End if the score is at least 25
            if score >= 25:
                break

    # If the best it found had at least 20
    if best_score >= 20:
        return best
    else:
        # If it found not results, search again without year
        if use_year and year is not None:
            return search_tmdb_match(db_session, show_data, use_year=False)

        print_message('no TMDB match found', True, str(show_data.id))
        return None


def add_file_data(db_session: sqlalchemy.orm.Session, channel_set: int, filename: str) -> ():
    """
    Select the function according to the channel set.

    :param db_session: the DB session.
    :param channel_set: the set of channels of the file.
    :param filename: the name of the file.
    """

    print('Processing file...')

    result = channel_insertion_list[channel_set].add_file_data(db_session, filename)

    if result is not None:
        print('complete!\n')
        print('The file contained %d show sessions!' % result.total_nb_sessions_in_file)
        print('Shows\' interval from %s to %s.\n' % (str(result.start_datetime), str(result.end_datetime)))

        print('%4d show sessions updated!' % result.nb_updated_sessions)
        print('%4d show sessions added!' % result.nb_added_sessions)
        print('%4d show sessions deleted!' % result.nb_deleted_sessions)
        print('%4d new shows!' % result.nb_new_shows)


def execute_data_insertion():
    """ Execute a data insertion. """

    question = 'Choose one channel set for the data being inserted:\n'

    for i in range(len(channel_insertion_list)):
        question += '%d - %s\n' % (i, channel_insertion_list[i].channels)

    input_channel_set = int(input(question))

    input_filename = input('What is the path to the file?\n')

    session = configuration.Session()

    try:
        add_file_data(session, input_channel_set, input_filename)
        session.commit()
    except:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == '__main__':
    execute_data_insertion()
